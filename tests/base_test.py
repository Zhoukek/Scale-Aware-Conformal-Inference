import os, sys
sys.path.insert(1, os.path.join(sys.path[0], '../'))
import numpy as np
import pandas as pd
from core import standard_weighted_quantile, trailing_window, aci, aci_clipped, quantile, quantile_integrator_log, quantile_integrator_log_scorecaster, ECI, OGD, SF_OGD, ECI_cutoff, ECI_integral, full_smoothed_eci, decay_OGD, smoothed_ogd, LQT, COP, SACI
from core.synthetic_scores import generate_scores
from core.model_scores import generate_forecasts
from datasets import load_dataset
from darts import TimeSeries
import yaml
import pickle
import pdb
import matplotlib.pyplot as plt
import logging
from scipy.stats import norm
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
import random
import torch

def set_all_seeds(seed):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


if __name__ == "__main__":
    json_name = sys.argv[1]
    seed = sys.argv[2]
    if len(sys.argv) > 2:
        overwrite = sys.argv[2].split(",")
    else:
        overwrite = []
    args = yaml.safe_load(open(json_name))
    # Set up folder and filename
    set_all_seeds(int(seed))
    seed = 'SEED_{}'.format(seed)
    foldername = './results/{}/'.format(seed) 
    config_name = json_name.split('.')[-2].split('/')[-1]
    filename = foldername + config_name + ".pkl"
    logname = foldername + config_name + ".log"
    os.makedirs(foldername, exist_ok=True)
    real_data = args['real']
    quantiles_given = args['quantiles_given']
    multiple_series = args['multiple_series']
    score_function_name = args['score_function_name'] if real_data else "synthetic"
    model_names = args['sequences'][0]['model_names'] if real_data else ["synthetic"]
    ahead = args['ahead'] if real_data and 'ahead' in args.keys() else 1
    minsize = args['minsize'] if real_data and 'minsize' in args.keys() else 0
    asymmetric = False

    logging.basicConfig(
        filename=logname,   
        level=logging.INFO,      
        format='%(asctime)s - %(levelname)s - %(message)s', 
        filemode='w'
    )
    
    all_results = {}
    # save the result to the Excel file path
    excel_file = foldername + '/{}_temp.xlsx'.format(config_name)
    if os.path.exists(excel_file):
        os.remove(excel_file)

    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        wb = Workbook()
        ws = wb.active
        ws.append(['Model name', 'Method', 'Average (%)', 'Average Interval (min)', 'Median Interval (min)'])
        wb.save(excel_file)


    for model_name in model_names:
        try:
            results = all_results[model_name]
        except:
            results = {}
        # Initialize the score function
        if real_data:
            score_function_name = args['score_function_name']
            if score_function_name == "absolute-residual":
                def score_function(y, forecast):
                    return np.abs(y - forecast)
                def set_function(forecast, q):
                    return np.array([forecast - q, forecast + q])
            elif score_function_name == "signed-residual":
                def score_function(y, forecast):
                    return np.array([forecast - y, y - forecast])
                def set_function(forecast, q):
                    return np.array([forecast - q[0], forecast + q[1]])
                asymmetric = True
            elif score_function_name == "cqr-symmetric":
                def score_function(y, forecasts):
                    return np.maximum(forecasts[0] - y, y - forecasts[-1])
                def set_function(forecast, q):
                    return np.array([forecast[0] - q, forecast[-1] + q])
            elif score_function_name == "cqr-asymmetric":
                def score_function(y, forecasts):
                    return np.array([forecasts[0] - y, y - forecasts[-1]])
                def set_function(forecast, q):
                    return np.array([forecast[0] - q[0], forecast[-1] + q[1]])
                asymmetric = True
            else:
                raise ValueError("Invalid score function name")

        # Get dataframe and add forecasts and scores to it
        if real_data:
            data = load_dataset(args['sequences'][0]['dataset'])
            # Get the forecasts
            if 'forecasts' not in data.columns:
                os.makedirs('./datasets/proc_{}/'.format(seed), exist_ok=True)
                os.makedirs('./datasets/proc_{}/'.format(seed) + config_name, exist_ok=True)
                args['sequences'][0]['savename'] = './datasets/proc_{}/'.format(seed) + config_name +  '/' + model_name + '.npz'
                args['sequences'][0]['T_burnin'] = args['T_burnin']
                args['sequences'][0]['ahead'] = ahead
                args['sequences'][0]['model_name'] = model_name
                args['sequences'][0]['output_chunk_length'] = args['sequences'][0]['fit_every']
                data['forecasts'] = generate_forecasts(data, **args['sequences'][0])
            # Compute scores
            if 'scores' not in data.columns:
                data['scores'] = [ score_function(y, forecast) for y, forecast in zip(data['y'], data['forecasts']) ] 
        else:
            scores_list = []
            for key in args['sequences'].keys():
                scores_list += [generate_scores(**args['sequences'][key])]
            scores = np.concatenate(scores_list).astype(float)
            # Make a pandas dataframe with a datetime index and the scores in their own column called `scores'.
            data = pd.DataFrame({'scores': scores}, index=pd.date_range(start='1/1/2018', periods=len(scores), freq='D'))

        # Loop through each method and learning rate, and compute the results
        for method in args['methods'].keys():
            if (method in results.keys()) and (method not in overwrite):
                continue
            fn = None
            if method == "Trail":
                fn = trailing_window
                args['methods'][method]['lrs'] = [None]
            elif method == "ACI":
                fn = aci
            elif method == "ACI (clipped)":
                fn = aci_clipped
            elif method == "OGD":
                fn = OGD
            elif method == "SF_OGD":
                fn = SF_OGD
            elif method == "decay_OGD":
                fn = decay_OGD
            elif method == "Quantile":
                fn = quantile
            elif method == "Quantile+Integrator (log)":
                fn = quantile_integrator_log
            elif method == "Quantile+Integrator (log)+Scorecaster":
                fn = quantile_integrator_log_scorecaster
            elif method == "ECI":
                fn = ECI
            elif method == "ECI_cutoff":
                fn = ECI_cutoff
            elif method == "ECI_integral":
                fn = ECI_integral
            elif method == 'full_smoothed_eci':
                fn = full_smoothed_eci
            elif method == 'smoothed_ogd':
                fn = smoothed_ogd
            elif method == 'LQT':
                fn = LQT
            elif method == 'COP':
                fn = COP
            elif method == 'SACI':
                fn = SACI
            else:
                raise Exception(f"Method {method} not implemented")
            lrs = args['methods'][method]['lrs']
            kwargs = args['methods'][method]
            kwargs["T_burnin"] = args["T_burnin"]
            kwargs["data"] = data if real_data else None
            kwargs["seasonal_period"] = args["seasonal_period"] if "seasonal_period" in args.keys() else None
            kwargs["config_name"] = config_name
            kwargs["ahead"] = ahead
            kwargs["seed"] = seed
            # Compute the results
            results[method] = {}
            # pdb.set_trace()
            average_interval_list = []
            median_interval_list = []
            cover_label_list = []
            for lr in lrs:
                if asymmetric:
                    stacked_scores = np.stack(data['scores'].to_list())
                    kwargs['upper'] = False
                    q0 = fn(stacked_scores[:,0], args['alpha']/2, lr, **kwargs)['q']
                    kwargs['upper'] = True
                    q1 = fn(stacked_scores[:,1], args['alpha']/2, lr, **kwargs)['q']
                    q = [ np.array([q0[i], q1[i]]) for i in range(len(q0)) ]
                else:
                    kwargs['upper'] = True
                    q = fn(data['scores'].to_numpy(), args['alpha'], lr, **kwargs)['q']
                if real_data:
                    sets = [ set_function(data['forecasts'].interpolate().to_numpy()[i], q[i]) for i in range(len(q)) ] # 获得预测完的q之后再去构建set
                    # Make sure the set size is at least minsize by setting sets[j][0] = min(sets[j][0], sets[j][1]-minsize) and sets[j][1] = max(sets[j][1], sets[j][1]+minsize)
                    sets = [ np.array([np.minimum(sets[j][0], sets[j][1]-minsize), np.maximum(sets[j][1], sets[j][0]+minsize)]) for j in range(len(sets)) ]
                else:
                    sets = None
                results[method][lr] = { "q": q, "sets": sets }
                
                T_burnin = args['T_burnin']
                if real_data:
                    # pdb.set_trace()
                    results[method][lr]["covered"] = (
                              (np.stack(results[method][lr]["sets"][T_burnin+1:])[:,0] <= data['y'][T_burnin+1:]) \
                            & (np.stack(results[method][lr]["sets"][T_burnin+1:])[:,1] >= data['y'][T_burnin+1:]) \
                        )
                else:
                    results[method][lr]["covered"] = results[method][lr]["q"][T_burnin+1:] >= scores[T_burnin+1:]
                coverages = {}
                coverages[method] = { lr : results[method][lr]["covered"].astype(int).mean() for lr in list(results[method].keys()) }

                label = f"cvg={100*coverages[method][lr]:.1f}%" if lr is not None else f"cvg={100*coverages[method][lr]:.1f}%"


                # print interval length
                intervals = [y - x for x, y in sets]
                average_interval = np.mean(intervals)
                median_interval = np.median(intervals)
                print(f"{model_name}: {method}: Average interval length={average_interval}, Median interval length={median_interval}")
                logging.info("{:<10} | {:<15} | lr={:>8.4f}, Avg={:>6.3f}, Med={:>6.3f}, ".format(model_name, method, lr, average_interval, median_interval, label) + label)
                cover_label_list.append(coverages[method][lr])
                average_interval_list.append(average_interval)
                median_interval_list.append(median_interval)
            
            for i in range(len(cover_label_list)):
                temp = float(f"{100 * cover_label_list[i]:.1f}")
                if 90.00 - temp > 1.000001:
                    average_interval_list[i] = 99999999
                    median_interval_list[i] = 99999999

            wb = load_workbook(excel_file)
            ws = wb.active
            
            if config_name in ['elec2_test']: 
                new_row = [
                    model_name,
                    method,
                    "{:.1f}".format(100 * cover_label_list[np.argmin(median_interval_list)]),
                    "{:.3f}".format(average_interval_list[np.argmin(median_interval_list)]),
                    "{:.3f}".format(min(median_interval_list))
                ]
            else:
                new_row = [
                    model_name,
                    method,
                    "{:.1f}".format(100 * cover_label_list[np.argmin(median_interval_list)]),
                    "{:.2f}".format(average_interval_list[np.argmin(median_interval_list)]),
                    "{:.2f}".format(min(median_interval_list))
                ]
    
            ws.append(new_row)
            wb.save(excel_file)

            logging.info(" average={:>6.1f}%  ,average_interval_min={:>6.3f}, median_interval_min={:>6.3f} \n".format(100*cover_label_list[np.argmin(average_interval_list)], average_interval_list[np.argmin(median_interval_list)], min(median_interval_list)))
        
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append(["", "", "", "", ""])
        wb.save(excel_file)

        # Save some metadata
        results["scores"] = data['scores']
        results["alpha"] = args['alpha']
        results["T_burnin"] = args['T_burnin']
        results["quantiles_given"] = quantiles_given
        results["multiple_series"] = multiple_series
        results["real_data"] = real_data
        results["score_function_name"] = score_function_name
        results["asymmetric"] = asymmetric

        if real_data:
            results["forecasts"] = data['forecasts']
            results["data"] = data
        all_results[model_name] = results

    # Save results
    with open(filename, 'wb') as handle:
        pickle.dump(all_results, handle, protocol=pickle.HIGHEST_PROTOCOL)

# test